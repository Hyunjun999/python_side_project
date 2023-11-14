# pip install openpyxl python-docx docx2pdf
import pandas as pd

# Each column means name - birth - number
df = pd.DataFrame(
    [
        ["John Doe", "1990.10.03", "2023-0001"],
        ["Jane Doe", "1993.05.05", "2023-0002"],
        ["Joe Smith", "2002.06.19", "2023-0003"],
        ["Sophia Smith", "2003. 07.31", "2023-0004"],
        ["Cave man", "2010.11.11", "2023-0005"],
        ["김철수", "2011.01.01", "2023-0006"],
        ["홍길동", "2009.12.14", "2023-0007"],
    ]
)
print(df)
df.to_excel(r"certificate_name_list.xlsx", index=False, header=False)

from openpyxl import load_workbook

load_wb = load_workbook(r"certificate_name_list.xlsx")
load_ws = load_wb.active

name, birth, num = [], [], []
for i in range(1, load_ws.max_row + 1):
    name.append(load_ws.cell(i, 1).value)
    birth.append(load_ws.cell(i, 2).value)
    num.append(load_ws.cell(i, 3).value)

print(name, birth, num)
