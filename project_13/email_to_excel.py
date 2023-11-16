import re
import requests
from openpyxl import load_workbook, Workbook

# url = "https://news.v.daum.net/v/20211129144552297"
# url = "https://www.washingtonpost.com/business/2023/11/15/senate-vote-avert-government-shutdown/"
url = [
    "https://news.v.daum.net/v/20211129144552297",
    "https://www.washingtonpost.com/business/2023/11/15/senate-vote-avert-government-shutdown/",
]
headers = {"User-Agent": "Mozilla/5.0", "Content-Type": "text/html; charset=utf-8"}
for u in url:
    response = requests.get(u, headers=headers)

    test_str = """
    aaa@bbb.com
    0123@abc.co.kr
    test@hello.kr
    ok@ok.co.kr
    ok@ok.co.kr
    ok@ok.co.kr
    no.co.kr
    no.kr
    """

    res = re.findall(
        r"[a-zA-Z0-9_\.-]+@[\w\.-]+", test_str
    )  # [\w] is equivalent to [a-zA-Z0-9_]
    res = re.findall(r"[a-zA-Z0-9_\.-]+@[\w\.-]+", response.text)
    res = set(res)
    print(res)

    try:
        wb = load_workbook(r"email.xlsx", data_only=True)
        sheet = wb.active
    except:
        wb = Workbook()
        sheet = wb.active

    for r in res:
        sheet.append([r])
    wb.save("email.xlsx")
